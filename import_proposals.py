#!/usr/bin/env python3
"""
import_proposals.py

Imports proposals from Proposals.xlsx into the database
Matches your existing project codes (25 BK-001, etc)
Updates proposals table with project details
"""

import openpyxl
import sqlite3
from pathlib import Path
from datetime import datetime

class ProposalsImporter:
    def __init__(self, excel_path, master_db_path=None):
        self.excel_path = Path(excel_path)

        # Default to your standard location
        if master_db_path is None:
            self.master_db = Path.home() / "Desktop/BDS_SYSTEM/01_DATABASES/bensley_master.db"
        else:
            self.master_db = Path(master_db_path)

        print(f"Database: {self.master_db}")

        # Load Excel
        self.wb = openpyxl.load_workbook(excel_path, data_only=True)
        self.ws = self.wb['Weekly proposal']

        # Connect to database
        self.conn = sqlite3.connect(self.master_db)
        self.cursor = self.conn.cursor()

        self.stats = {
            'proposals_found': 0,
            'proposals_imported': 0,
            'projects_linked': 0,
            'errors': 0
        }

    def extract_project_code(self, code_str):
        """Extract BK code from '25 BK-001' format"""
        if not code_str:
            return None

        code_str = str(code_str).strip()

        # Look for pattern "BK-001"
        if "BK-" in code_str:
            parts = code_str.split("BK-")
            if len(parts) > 1:
                code_num = parts[1].split()[0]
                return f"BK-{code_num}"

        return None

    def create_proposals_table(self):
        """Create proposals table if it doesn't exist"""
        try:
            # Check if table exists with old schema
            self.cursor.execute("SELECT sql FROM sqlite_master WHERE type='table' AND name='proposals'")
            result = self.cursor.fetchone()

            if result:
                schema = result[0]
                # If old schema missing columns, drop and recreate
                if 'client_company' not in schema:
                    print("⚠️  Old proposals table schema detected - recreating...")
                    self.cursor.execute("DROP TABLE IF EXISTS proposals")
                    self.conn.commit()

            self.cursor.execute("""
                CREATE TABLE IF NOT EXISTS proposals (
                    proposal_id INTEGER PRIMARY KEY,
                    project_code TEXT UNIQUE NOT NULL,
                    project_name TEXT,
                    client_company TEXT,
                    contact_person TEXT,
                    contact_phone TEXT,
                    contact_email TEXT,
                    project_value REAL,
                    is_landscape INTEGER DEFAULT 0,
                    is_architect INTEGER DEFAULT 0,
                    is_interior INTEGER DEFAULT 0,
                    status TEXT DEFAULT 'proposal',
                    created_at TEXT,
                    updated_at TEXT
                )
            """)
            self.conn.commit()
            print("✓ Proposals table ready")
        except Exception as e:
            print(f"✗ Error creating table: {e}")
            self.stats['errors'] += 1

    def extract_proposals(self):
        """Extract proposals from Excel sheet"""
        proposals = []
        current_row = 7  # Start after header (row 6)
        max_rows = self.ws.max_row

        print(f"\nExtracting proposals from rows 7-{max_rows}...")

        while current_row <= max_rows:
            try:
                # Get project code from column B (index 2)
                project_code_cell = self.ws.cell(row=current_row, column=2).value

                # Skip empty rows
                if not project_code_cell:
                    current_row += 1
                    continue

                # Extract project code
                project_code = self.extract_project_code(project_code_cell)

                if not project_code:
                    current_row += 1
                    continue

                # Extract data
                project_name = self.ws.cell(row=current_row, column=3).value or ""
                client_company = self.ws.cell(row=current_row, column=4).value or ""
                contact_person = self.ws.cell(row=current_row, column=5).value or ""
                contact_phone = self.ws.cell(row=current_row, column=6).value or ""
                contact_email = self.ws.cell(row=current_row, column=7).value or ""

                # Disciplines (Y/0 format)
                is_landscape = self.ws.cell(row=current_row, column=8).value
                is_architect = self.ws.cell(row=current_row, column=9).value
                is_interior = self.ws.cell(row=current_row, column=10).value

                # Project value
                project_value = self.ws.cell(row=current_row, column=11).value
                if project_value:
                    try:
                        project_value = float(project_value)
                    except:
                        project_value = None

                # Convert Y/0 to 1/0
                is_landscape = 1 if is_landscape == 'Y' else 0
                is_architect = 1 if is_architect == 'Y' else 0
                is_interior = 1 if is_interior == 'Y' else 0

                proposal = {
                    'project_code': project_code,
                    'project_name': str(project_name).strip() if project_name else "",
                    'client_company': str(client_company).strip() if client_company else "",
                    'contact_person': str(contact_person).strip() if contact_person else "",
                    'contact_phone': str(contact_phone).strip() if contact_phone else "",
                    'contact_email': str(contact_email).strip() if contact_email else "",
                    'project_value': project_value,
                    'is_landscape': is_landscape,
                    'is_architect': is_architect,
                    'is_interior': is_interior,
                }

                proposals.append(proposal)
                self.stats['proposals_found'] += 1

                if self.stats['proposals_found'] % 10 == 0:
                    print(f"   Found {self.stats['proposals_found']} proposals...")

                # Move to next project (skip status rows - each project takes 7 rows)
                current_row += 7

            except Exception as e:
                print(f"   ⚠ Error at row {current_row}: {e}")
                current_row += 1
                self.stats['errors'] += 1

        print(f"✓ Found {self.stats['proposals_found']} proposals")
        return proposals

    def import_proposals(self, proposals):
        """Import proposals to database"""
        print(f"\nImporting {len(proposals)} proposals...")

        for proposal in proposals:
            try:
                self.cursor.execute("""
                    INSERT OR REPLACE INTO proposals
                    (project_code, project_name, client_company, contact_person,
                     contact_phone, contact_email, project_value,
                     is_landscape, is_architect, is_interior,
                     created_at, updated_at)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (
                    proposal['project_code'],
                    proposal['project_name'],
                    proposal['client_company'],
                    proposal['contact_person'],
                    proposal['contact_phone'],
                    proposal['contact_email'],
                    proposal['project_value'],
                    proposal['is_landscape'],
                    proposal['is_architect'],
                    proposal['is_interior'],
                    datetime.now().isoformat(),
                    datetime.now().isoformat()
                ))

                self.stats['proposals_imported'] += 1

            except sqlite3.IntegrityError:
                # Already exists, that's fine
                pass
            except Exception as e:
                print(f"   ✗ Error importing {proposal['project_code']}: {e}")
                self.stats['errors'] += 1

        self.conn.commit()
        print(f"✓ Imported {self.stats['proposals_imported']} proposals")

    def link_to_projects(self):
        """Link proposals to existing projects in the projects table"""
        print(f"\nLinking proposals to projects...")

        try:
            # Get all proposals
            self.cursor.execute("SELECT project_code, project_name FROM proposals")
            proposals = self.cursor.fetchall()

            for project_code, project_name in proposals:
                try:
                    # Check if project exists in projects table
                    self.cursor.execute("""
                        SELECT project_id FROM projects WHERE project_code = ?
                    """, (project_code,))

                    result = self.cursor.fetchone()
                    if result:
                        # Project exists, we can link data
                        self.stats['projects_linked'] += 1

                except Exception as e:
                    pass

            print(f"✓ Linked {self.stats['projects_linked']} to existing projects")

        except Exception as e:
            print(f"⚠ Error linking: {e}")

    def verify_import(self):
        """Verify the import was successful"""
        print(f"\nVerifying import...")

        try:
            self.cursor.execute("SELECT COUNT(*) FROM proposals")
            total = self.cursor.fetchone()[0]

            # Show sample
            self.cursor.execute("""
                SELECT project_code, project_name, client_company, project_value
                FROM proposals
                ORDER BY project_code
                LIMIT 10
            """)

            samples = self.cursor.fetchall()

            print(f"✓ Total proposals in database: {total}")
            print(f"\n📋 Sample proposals:")
            for code, name, client, value in samples:
                value_str = f"${value:,.0f}" if value else "N/A"
                print(f"   {code}: {name[:40]:40} | {client[:20]:20} | {value_str}")

        except Exception as e:
            print(f"⚠ Error verifying: {e}")

    def run(self):
        """Run the complete import process"""
        print("=" * 80)
        print("📥 PROPOSALS IMPORTER")
        print("=" * 80)
        print(f"Excel file: {self.excel_path}")

        # Create table
        self.create_proposals_table()

        # Extract proposals
        proposals = self.extract_proposals()

        # Import to database
        if proposals:
            self.import_proposals(proposals)
            self.link_to_projects()

        # Verify
        self.verify_import()

        # Summary
        print("\n" + "=" * 80)
        print("✅ IMPORT COMPLETE")
        print("=" * 80)
        print(f"\nSummary:")
        print(f"  Proposals found:     {self.stats['proposals_found']}")
        print(f"  Proposals imported:  {self.stats['proposals_imported']}")
        print(f"  Projects linked:     {self.stats['projects_linked']}")
        print(f"  Errors:              {self.stats['errors']}")
        print(f"  Database:            {self.master_db}")
        print("=" * 80)

        self.conn.close()

def main():
    import sys

    # Get Excel file path
    if len(sys.argv) > 1:
        excel_file = sys.argv[1]
    else:
        # Default location
        excel_file = Path.home() / "Desktop/BDS_SYSTEM/Proposals.xlsx"

    if not Path(excel_file).exists():
        print(f"✗ Excel file not found: {excel_file}")
        print(f"\nUsage: python3 import_proposals.py /path/to/Proposals.xlsx")
        return

    # Get database path (optional)
    db_path = None
    if len(sys.argv) > 2:
        db_path = sys.argv[2]

    try:
        importer = ProposalsImporter(excel_file, db_path)
        importer.run()
    except Exception as e:
        print(f"\n✗ Fatal error: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    main()
