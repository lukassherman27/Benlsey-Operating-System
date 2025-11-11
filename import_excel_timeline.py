#!/usr/bin/env python3
"""
import_excel_timeline.py

Imports real project dates from your Excel weekly proposal tracking sheet
Updates database with:
- First Contact date
- Proposal Sent date  
- Contract Signed date
- Current status
"""

import openpyxl
import sqlite3
from pathlib import Path
from datetime import datetime

class ExcelTimelineImporter:
    def __init__(self, excel_path):
        self.excel_path = excel_path
        self.master_db = Path.home() / "Desktop/BDS_SYSTEM/01_DATABASES/bensley_master.db"
        
        # Load Excel with calculated values
        self.wb = openpyxl.load_workbook(excel_path, data_only=True)
        self.ws = self.wb['Weekly proposal']
        
        # Database
        self.conn = sqlite3.connect(self.master_db)
        self.cursor = self.conn.cursor()
        
        self.stats = {
            'projects_found': 0,
            'projects_updated': 0,
            'projects_not_found': 0,
            'dates_imported': 0
        }
    
    def extract_project_timeline(self, project_row):
        """Extract timeline data for a project by finding status changes in the consolidated row"""
        project_code = self.ws.cell(project_row, 2).value  # Column B
        if not project_code:
            return None
        
        project_data = {
            'project_code': str(project_code).strip(),
            'project_name': self.ws.cell(project_row, 3).value,
            'contact_company': self.ws.cell(project_row, 4).value,
            'contact_person': self.ws.cell(project_row, 5).value,
            'contact_email': self.ws.cell(project_row, 7).value,
            'project_value': self.ws.cell(project_row, 11).value,
            'timeline': {}
        }
        
        # Read the consolidated status row - it shows current status for each date
        # Scan columns O onwards (col 15+) where dates are
        prev_status = None
        
        for col in range(15, 200):  # Column O onwards (dates)
            try:
                # Get the date from header row (row 6)
                date_cell = self.ws.cell(6, col)
                if not date_cell.value or not isinstance(date_cell.value, datetime):
                    continue
                
                date_str = date_cell.value.strftime('%Y-%m-%d')
                
                # Get the status for this date from the project row
                status_cell = self.ws.cell(project_row, col)
                current_status = status_cell.value
                
                # If status changed, record the date
                if current_status and current_status != prev_status:
                    # Clean up status name
                    status_clean = str(current_status).strip()
                    
                    # Only record if we don't already have this status (use earliest)
                    if status_clean not in project_data['timeline']:
                        project_data['timeline'][status_clean] = date_str
                    
                    prev_status = current_status
                
            except Exception as e:
                # End of data
                break
        
        return project_data
    
    def import_all_projects(self):
        """Import all projects from Excel"""
        print("="*70)
        print("IMPORTING PROJECT TIMELINES FROM EXCEL")
        print("="*70)
        print(f"\nReading: {self.excel_path}")
        
        # Find header row
        header_row = None
        for row in range(1, 20):
            if self.ws.cell(row, 2).value and 'Project No' in str(self.ws.cell(row, 2).value):
                header_row = row
                break
        
        if not header_row:
            print("❌ Could not find header row!")
            return
        
        print(f"Header found at row {header_row}")
        print("\nExtracting projects...")
        
        # Extract each project (they appear every ~7 rows)
        current_row = header_row + 1
        projects_data = []
        
        while current_row < self.ws.max_row:
            cell = self.ws.cell(current_row, 2)
            
            # Check if this is a project row (has a project code like "25 BK-001")
            if cell.value and isinstance(cell.value, str) and 'BK-' in cell.value:
                project = self.extract_project_timeline(current_row)
                if project:
                    projects_data.append(project)
                    self.stats['projects_found'] += 1
                
                # Move to next project (skip status rows)
                current_row += 7
            else:
                current_row += 1
        
        print(f"✅ Found {len(projects_data)} projects in Excel\n")
        
        # Update database
        print("Updating database...")
        for proj in projects_data:
            self.update_project_in_db(proj)
        
        self.conn.commit()
        
        # Show summary
        print("\n" + "="*70)
        print("IMPORT COMPLETE")
        print("="*70)
        print(f"\n✅ Projects found in Excel: {self.stats['projects_found']}")
        print(f"✅ Projects updated in DB: {self.stats['projects_updated']}")
        print(f"✅ Total dates imported: {self.stats['dates_imported']}")
        print(f"⚠️  Projects not found in DB: {self.stats['projects_not_found']}")
        
        # Show sample updates
        print("\n📊 Sample imported data:")
        self.show_sample_updates()
        
        self.conn.close()
    
    def update_project_in_db(self, proj_data):
        """Update a project in the database with timeline data"""
        project_code = proj_data['project_code']
        
        # Find project in database
        self.cursor.execute("""
            SELECT project_id FROM projects WHERE project_code = ?
        """, (project_code,))
        
        result = self.cursor.fetchone()
        if not result:
            self.stats['projects_not_found'] += 1
            return
        
        project_id = result[0]
        
        # Update project with first contact date (if available)
        timeline = proj_data['timeline']
        if 'First Contact' in timeline:
            first_contact = timeline['First Contact']
            
            self.cursor.execute("""
                UPDATE projects 
                SET date_created = ?
                WHERE project_id = ?
            """, (first_contact, project_id))
            
            self.stats['dates_imported'] += 1
        
        # Store full timeline in project_metadata
        for status, date in timeline.items():
            try:
                self.cursor.execute("""
                    INSERT OR REPLACE INTO project_metadata
                    (project_id, metadata_key, metadata_value, source, confidence)
                    VALUES (?, ?, ?, 'excel_import', 1.0)
                """, (project_id, f'timeline_{status.lower().replace(" ", "_")}', date))
                
                self.stats['dates_imported'] += 1
            except:
                pass
        
        # Update project value if available
        if proj_data.get('project_value'):
            try:
                value = float(proj_data['project_value'])
                self.cursor.execute("""
                    UPDATE projects
                    SET total_fee_usd = ?
                    WHERE project_id = ?
                """, (value, project_id))
            except:
                pass
        
        self.stats['projects_updated'] += 1
    
    def show_sample_updates(self):
        """Show some examples of imported data"""
        self.cursor.execute("""
            SELECT 
                p.project_code,
                p.project_title,
                p.date_created,
                COUNT(pm.metadata_id) as timeline_events
            FROM projects p
            LEFT JOIN project_metadata pm ON p.project_id = pm.project_id
                AND pm.metadata_key LIKE 'timeline_%'
            WHERE p.date_created < '2025-10-22'  -- Exclude fake dates
            GROUP BY p.project_id
            ORDER BY p.date_created DESC
            LIMIT 5
        """)
        
        for code, title, date, events in self.cursor.fetchall():
            print(f"\n   {code} - {title[:40]}")
            print(f"      First contact: {date}")
            print(f"      Timeline events: {events}")

def main():
    excel_file = Path.home() / "Desktop/BDS_SYSTEM/Proposals_overview_.xlsx"
    
    if not excel_file.exists():
        print(f"❌ Excel file not found: {excel_file}")
        print("\nPlease copy Proposals_overview_.xlsx to:")
        print(f"   {Path.home() / 'Desktop/BDS_SYSTEM/'}")
        return
    
    importer = ExcelTimelineImporter(excel_file)
    importer.import_all_projects()

if __name__ == '__main__':
    main()
