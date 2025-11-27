#!/usr/bin/env python3
"""
Create final formatted Excel file from parsed invoice data
"""
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Read the parsed CSV
csv_file = "reports/ALL_INVOICES_PARSED.csv"
df = pd.read_csv(csv_file)

# Sort by project_code, discipline, phase, invoice_date
df = df.sort_values(['project_code', 'discipline', 'phase', 'invoice_date'], na_position='last')

# Create Excel writer
excel_file = "/Users/lukassherman/Desktop/COMPLETE_INVOICES_FOR_AUDIT.xlsx"
writer = pd.ExcelWriter(excel_file, engine='openpyxl')

# Write main sheet
df.to_excel(writer, sheet_name='All Invoices', index=False)

# Get workbook and worksheet
workbook = writer.book
worksheet = writer.sheets['All Invoices']

# Define styles
header_fill = PatternFill(start_color="2C5F2D", end_color="2C5F2D", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)
outstanding_fill = PatternFill(start_color="FFF4E6", end_color="FFF4E6", fill_type="solid")
partial_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
paid_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")

border = Border(
    left=Side(style='thin', color='CCCCCC'),
    right=Side(style='thin', color='CCCCCC'),
    top=Side(style='thin', color='CCCCCC'),
    bottom=Side(style='thin', color='CCCCCC')
)

# Format header row
for cell in worksheet[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = border

# Format data rows
for row_idx in range(2, worksheet.max_row + 1):
    status_cell = worksheet[f'L{row_idx}']  # status column
    status = status_cell.value

    # Apply row coloring
    for col_idx in range(1, worksheet.max_column + 1):
        cell = worksheet.cell(row=row_idx, column=col_idx)
        cell.border = border
        cell.alignment = Alignment(vertical='center')

        # Color based on status
        if status == 'outstanding':
            cell.fill = outstanding_fill
        elif status == 'partial':
            cell.fill = partial_fill
        elif status == 'paid':
            cell.fill = paid_fill

# Adjust column widths
column_widths = {
    'A': 12,  # project_code
    'B': 45,  # project_name
    'C': 14,  # invoice_number
    'D': 12,  # invoice_date
    'E': 15,  # invoice_amount
    'F': 12,  # payment_date
    'G': 15,  # payment_amount
    'H': 15,  # outstanding_amount
    'I': 25,  # phase
    'J': 22,  # discipline
    'K': 20,  # notes
    'L': 12,  # status
}

for col_letter, width in column_widths.items():
    worksheet.column_dimensions[col_letter].width = width

# Format currency columns
for row_idx in range(2, worksheet.max_row + 1):
    worksheet[f'E{row_idx}'].number_format = '$#,##0.00'
    worksheet[f'G{row_idx}'].number_format = '$#,##0.00'
    worksheet[f'H{row_idx}'].number_format = '$#,##0.00'

# Format date columns
for row_idx in range(2, worksheet.max_row + 1):
    if worksheet[f'D{row_idx}'].value:
        worksheet[f'D{row_idx}'].number_format = 'YYYY-MM-DD'
    if worksheet[f'F{row_idx}'].value:
        worksheet[f'F{row_idx}'].number_format = 'YYYY-MM-DD'

# Freeze header row
worksheet.freeze_panes = 'A2'

# Create summary sheet by project
summary_df = df.groupby(['project_code', 'project_name']).agg({
    'invoice_amount': 'sum',
    'payment_amount': 'sum',
    'outstanding_amount': 'sum',
    'invoice_number': 'count'
}).reset_index()

summary_df.columns = ['Project Code', 'Project Name', 'Total Invoiced', 'Total Paid', 'Total Outstanding', 'Invoice Count']
summary_df = summary_df.sort_values('Total Invoiced', ascending=False)

summary_df.to_excel(writer, sheet_name='Summary by Project', index=False)

# Format summary sheet
summary_sheet = writer.sheets['Summary by Project']

for cell in summary_sheet[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = border

for row in summary_sheet.iter_rows(min_row=2, max_row=summary_sheet.max_row):
    for cell in row:
        cell.border = border

# Adjust summary column widths
summary_sheet.column_dimensions['A'].width = 14
summary_sheet.column_dimensions['B'].width = 50
summary_sheet.column_dimensions['C'].width = 18
summary_sheet.column_dimensions['D'].width = 18
summary_sheet.column_dimensions['E'].width = 18
summary_sheet.column_dimensions['F'].width = 14

# Format currency in summary
for row_idx in range(2, summary_sheet.max_row + 1):
    summary_sheet[f'C{row_idx}'].number_format = '$#,##0.00'
    summary_sheet[f'D{row_idx}'].number_format = '$#,##0.00'
    summary_sheet[f'E{row_idx}'].number_format = '$#,##0.00'

summary_sheet.freeze_panes = 'A2'

# Create outstanding invoices sheet
outstanding_df = df[df['status'] == 'outstanding'].copy()
outstanding_df = outstanding_df.sort_values(['project_code', 'invoice_amount'], ascending=[True, False])
outstanding_df.to_excel(writer, sheet_name='Outstanding Invoices', index=False)

outstanding_sheet = writer.sheets['Outstanding Invoices']

# Format header
for cell in outstanding_sheet[1]:
    cell.fill = PatternFill(start_color="D32F2F", end_color="D32F2F", fill_type="solid")
    cell.font = Font(bold=True, color="FFFFFF", size=11)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = border

# Format rows
for row in outstanding_sheet.iter_rows(min_row=2, max_row=outstanding_sheet.max_row):
    for cell in row:
        cell.fill = outstanding_fill
        cell.border = border

# Adjust column widths
for col_letter, width in column_widths.items():
    outstanding_sheet.column_dimensions[col_letter].width = width

# Format currency and dates
for row_idx in range(2, outstanding_sheet.max_row + 1):
    outstanding_sheet[f'E{row_idx}'].number_format = '$#,##0.00'
    outstanding_sheet[f'G{row_idx}'].number_format = '$#,##0.00'
    outstanding_sheet[f'H{row_idx}'].number_format = '$#,##0.00'
    if outstanding_sheet[f'D{row_idx}'].value:
        outstanding_sheet[f'D{row_idx}'].number_format = 'YYYY-MM-DD'

outstanding_sheet.freeze_panes = 'A2'

# Create partial payments sheet
partial_df = df[df['status'] == 'partial'].copy()
if len(partial_df) > 0:
    partial_df.to_excel(writer, sheet_name='Partial Payments', index=False)

    partial_sheet = writer.sheets['Partial Payments']

    for cell in partial_sheet[1]:
        cell.fill = PatternFill(start_color="FF9800", end_color="FF9800", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border

    for row in partial_sheet.iter_rows(min_row=2, max_row=partial_sheet.max_row):
        for cell in row:
            cell.fill = partial_fill
            cell.border = border

    for col_letter, width in column_widths.items():
        partial_sheet.column_dimensions[col_letter].width = width

    for row_idx in range(2, partial_sheet.max_row + 1):
        partial_sheet[f'E{row_idx}'].number_format = '$#,##0.00'
        partial_sheet[f'G{row_idx}'].number_format = '$#,##0.00'
        partial_sheet[f'H{row_idx}'].number_format = '$#,##0.00'

    partial_sheet.freeze_panes = 'A2'

# Save
writer.close()

print("=" * 100)
print("EXCEL FILE CREATED FOR AUDIT")
print("=" * 100)
print(f"\nFile: {excel_file}")
print(f"\nStatistics:")
print(f"   Total Invoices: {len(df)}")
print(f"   Paid: {len(df[df['status'] == 'paid'])}")
print(f"   Partial: {len(df[df['status'] == 'partial'])}")
print(f"   Outstanding: {len(df[df['status'] == 'outstanding'])}")
print(f"\nFinancial Summary:")
print(f"   Total Invoiced: ${df['invoice_amount'].sum():,.2f}")
print(f"   Total Paid: ${df['payment_amount'].sum():,.2f}")
print(f"   Total Outstanding: ${df['outstanding_amount'].sum():,.2f}")
print(f"\nSheets Created:")
print(f"   1. All Invoices ({len(df)} invoices)")
print(f"   2. Summary by Project ({len(summary_df)} projects)")
print(f"   3. Outstanding Invoices ({len(outstanding_df)} invoices)")
if len(partial_df) > 0:
    print(f"   4. Partial Payments ({len(partial_df)} invoices)")
print(f"\nColor Coding:")
print(f"   🟢 Green = Paid")
print(f"   🟡 Yellow = Partial Payment")
print(f"   🟠 Orange = Outstanding")
print("\n✅ Ready for your audit!")
print("=" * 100)
