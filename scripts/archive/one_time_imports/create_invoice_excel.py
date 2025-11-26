#!/usr/bin/env python3
"""
Create a formatted Excel file from the parsed invoice CSV data
Fills in all columns and adds proper formatting for user audit
"""
import pandas as pd
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Read the CSV
csv_file = "reports/invoices_for_review.csv"
df = pd.read_csv(csv_file)

# Sort by project_code, then invoice_date
df = df.sort_values(['project_code', 'invoice_date'], na_position='last')

# Create Excel writer
excel_file = "/Users/lukassherman/Desktop/INVOICES_FOR_AUDIT.xlsx"
writer = pd.ExcelWriter(excel_file, engine='openpyxl')

# Write main sheet
df.to_excel(writer, sheet_name='All Invoices', index=False)

# Get the workbook and worksheet
workbook = writer.book
worksheet = writer.sheets['All Invoices']

# Define styles
header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)
outstanding_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
needs_review_fill = PatternFill(start_color="F4CCCC", end_color="F4CCCC", fill_type="solid")
paid_fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")

border = Border(
    left=Side(style='thin', color='000000'),
    right=Side(style='thin', color='000000'),
    top=Side(style='thin', color='000000'),
    bottom=Side(style='thin', color='000000')
)

# Format header row
for cell in worksheet[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = border

# Format data rows
for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, max_row=worksheet.max_row), start=2):
    # Get status and needs_review values
    status_cell = worksheet[f'K{row_idx}']  # status column
    needs_review_cell = worksheet[f'L{row_idx}']  # needs_review column

    status = status_cell.value
    needs_review = needs_review_cell.value

    # Apply row coloring
    for cell in row:
        cell.border = border
        cell.alignment = Alignment(vertical='center')

        # Color based on status
        if needs_review == 'YES':
            cell.fill = needs_review_fill
        elif status == 'outstanding':
            cell.fill = outstanding_fill
        elif status == 'paid':
            cell.fill = paid_fill

# Adjust column widths
column_widths = {
    'A': 15,  # project_code
    'B': 40,  # project_name
    'C': 15,  # invoice_number
    'D': 12,  # invoice_date
    'E': 15,  # invoice_amount
    'F': 12,  # payment_date
    'G': 15,  # payment_amount
    'H': 15,  # phase
    'I': 15,  # discipline
    'J': 30,  # notes
    'K': 12,  # status
    'L': 12,  # needs_review
}

for col_letter, width in column_widths.items():
    worksheet.column_dimensions[col_letter].width = width

# Format currency columns
for row_idx in range(2, worksheet.max_row + 1):
    worksheet[f'E{row_idx}'].number_format = '$#,##0.00'
    worksheet[f'G{row_idx}'].number_format = '$#,##0.00'

# Format date columns
for row_idx in range(2, worksheet.max_row + 1):
    worksheet[f'D{row_idx}'].number_format = 'YYYY-MM-DD'
    worksheet[f'F{row_idx}'].number_format = 'YYYY-MM-DD'

# Freeze header row
worksheet.freeze_panes = 'A2'

# Create summary sheet by project
summary_df = df.groupby(['project_code', 'project_name']).agg({
    'invoice_amount': 'sum',
    'payment_amount': 'sum',
    'invoice_number': 'count'
}).reset_index()

summary_df.columns = ['Project Code', 'Project Name', 'Total Invoiced', 'Total Paid', 'Invoice Count']
summary_df['Outstanding'] = summary_df['Total Invoiced'] - summary_df['Total Paid']

# Add summary for rows with no project code
no_project_df = df[df['project_code'].isna() | (df['project_code'] == '')]
if len(no_project_df) > 0:
    no_project_row = {
        'Project Code': 'UNKNOWN',
        'Project Name': 'Missing Project Code - NEEDS REVIEW',
        'Total Invoiced': no_project_df['invoice_amount'].sum(),
        'Total Paid': no_project_df['payment_amount'].sum(),
        'Invoice Count': len(no_project_df),
        'Outstanding': no_project_df['invoice_amount'].sum() - no_project_df['payment_amount'].sum()
    }
    summary_df = pd.concat([pd.DataFrame([no_project_row]), summary_df], ignore_index=True)

summary_df.to_excel(writer, sheet_name='Summary by Project', index=False)

# Format summary sheet
summary_sheet = writer.sheets['Summary by Project']

for cell in summary_sheet[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = border

for row in summary_sheet.iter_rows(min_row=2, max_row=summary_sheet.max_row):
    for cell in row:
        cell.border = border

# Adjust summary column widths
summary_sheet.column_dimensions['A'].width = 15
summary_sheet.column_dimensions['B'].width = 50
summary_sheet.column_dimensions['C'].width = 18
summary_sheet.column_dimensions['D'].width = 18
summary_sheet.column_dimensions['E'].width = 15
summary_sheet.column_dimensions['F'].width = 18

# Format currency in summary
for row_idx in range(2, summary_sheet.max_row + 1):
    summary_sheet[f'C{row_idx}'].number_format = '$#,##0.00'
    summary_sheet[f'D{row_idx}'].number_format = '$#,##0.00'
    summary_sheet[f'F{row_idx}'].number_format = '$#,##0.00'

    # Highlight UNKNOWN row
    if summary_sheet[f'A{row_idx}'].value == 'UNKNOWN':
        for col in ['A', 'B', 'C', 'D', 'E', 'F']:
            summary_sheet[f'{col}{row_idx}'].fill = needs_review_fill

summary_sheet.freeze_panes = 'A2'

# Create "Needs Review" sheet - only invoices that need attention
needs_review_df = df[df['needs_review'] == 'YES'].copy()
if len(needs_review_df) > 0:
    needs_review_df.to_excel(writer, sheet_name='NEEDS REVIEW', index=False)

    review_sheet = writer.sheets['NEEDS REVIEW']

    # Format header
    for cell in review_sheet[1]:
        cell.fill = PatternFill(start_color="CC0000", end_color="CC0000", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF", size=12)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border

    # Format rows
    for row in review_sheet.iter_rows(min_row=2, max_row=review_sheet.max_row):
        for cell in row:
            cell.fill = needs_review_fill
            cell.border = border

    # Adjust column widths
    for col_letter, width in column_widths.items():
        review_sheet.column_dimensions[col_letter].width = width

    # Format currency and dates
    for row_idx in range(2, review_sheet.max_row + 1):
        review_sheet[f'E{row_idx}'].number_format = '$#,##0.00'
        review_sheet[f'G{row_idx}'].number_format = '$#,##0.00'
        review_sheet[f'D{row_idx}'].number_format = 'YYYY-MM-DD'
        review_sheet[f'F{row_idx}'].number_format = 'YYYY-MM-DD'

    review_sheet.freeze_panes = 'A2'

# Create "Outstanding" sheet - only unpaid invoices
outstanding_df = df[df['status'] == 'outstanding'].copy()
outstanding_df = outstanding_df.sort_values(['invoice_amount'], ascending=False)
outstanding_df.to_excel(writer, sheet_name='Outstanding Invoices', index=False)

outstanding_sheet = writer.sheets['Outstanding Invoices']

# Format header
for cell in outstanding_sheet[1]:
    cell.fill = PatternFill(start_color="F9CB9C", end_color="F9CB9C", fill_type="solid")
    cell.font = Font(bold=True, size=11)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = border

# Format rows
for row in outstanding_sheet.iter_rows(min_row=2, max_row=outstanding_sheet.max_row):
    for cell in row:
        cell.fill = outstanding_fill
        cell.border = border

# Adjust column widths
for col_letter, width in column_widths.items():
    outstanding_sheet.column_dimensions[col_letter].width = width

# Format currency and dates
for row_idx in range(2, outstanding_sheet.max_row + 1):
    outstanding_sheet[f'E{row_idx}'].number_format = '$#,##0.00'
    outstanding_sheet[f'G{row_idx}'].number_format = '$#,##0.00'
    outstanding_sheet[f'D{row_idx}'].number_format = 'YYYY-MM-DD'
    outstanding_sheet[f'F{row_idx}'].number_format = 'YYYY-MM-DD'

outstanding_sheet.freeze_panes = 'A2'

# Save
writer.close()

print("=" * 80)
print("EXCEL FILE CREATED FOR AUDIT")
print("=" * 80)
print(f"\n📄 File: {excel_file}")
print(f"\n📊 Statistics:")
print(f"   Total Invoices: {len(df)}")
print(f"   Paid: {len(df[df['status'] == 'paid'])}")
print(f"   Outstanding: {len(df[df['status'] == 'outstanding'])}")
print(f"   Needs Review: {len(df[df['needs_review'] == 'YES'])}")
print(f"\n💰 Financial Summary:")
print(f"   Total Invoiced: ${df['invoice_amount'].sum():,.2f}")
print(f"   Total Paid: ${df['payment_amount'].sum():,.2f}")
print(f"   Outstanding: ${df[df['status'] == 'outstanding']['invoice_amount'].sum():,.2f}")
print(f"\n📑 Sheets Created:")
print(f"   1. All Invoices - Complete list with color coding")
print(f"   2. Summary by Project - Grouped totals")
print(f"   3. NEEDS REVIEW - {len(needs_review_df)} invoices missing project codes")
print(f"   4. Outstanding Invoices - {len(outstanding_df)} unpaid invoices")
print("\n🎨 Color Coding:")
print("   🔴 Red = Needs Review (missing project code)")
print("   🟡 Yellow = Outstanding (not yet paid)")
print("   🟢 Green = Paid")
print("\n✅ Ready for your audit!")
print("=" * 80)
