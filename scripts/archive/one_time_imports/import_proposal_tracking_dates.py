#!/usr/bin/env python3
"""
Import Proposal Tracking Dates from Proposals.xlsx Calendar Structure

Extracts historical tracking dates from your weekly proposal calendar:
- First Contact date
- Drafting date
- Proposal Sent date
- Contract Signed date

Updates the projects table with this historical data.
"""
import openpyxl
import sqlite3
from datetime import datetime
from pathlib import Path

EXCEL_FILE = "/Users/lukassherman/Library/CloudStorage/OneDrive-Personal/Bensley/Benlsey-Operating-System/Proposals.xlsx"
DB_PATH = "database/bensley_master.db"

def extract_project_code(code_str):
    """Extract BK code from '25 BK-001' format (keep full format with year)"""
    if not code_str:
        return None

    code_str = str(code_str).strip()

    # Look for pattern "YY BK-XXX" and keep the full format
    if "BK-" in code_str:
        return code_str  # Return full code like "25 BK-001"

    return None

def parse_date(date_val):
    """Convert datetime to YYYY-MM-DD format"""
    if date_val is None:
        return None

    if isinstance(date_val, datetime):
        return date_val.strftime("%Y-%m-%d")

    if isinstance(date_val, str):
        try:
            dt = datetime.strptime(date_val, "%Y-%m-%d %H:%M:%S")
            return dt.strftime("%Y-%m-%d")
        except:
            return None

    return None

def get_most_recent_marked_date(ws, row, start_col=15):
    """
    Scan across calendar columns (starting from col 15/O) to find most recent marked date.
    Returns the date from the column header if that column has any value marked.
    """
    max_date = None
    max_col = ws.max_column

    # Scan through all calendar columns
    for col in range(start_col, max_col + 1):
        # Check if this column has a value in the target row
        cell_value = ws.cell(row=row, column=col).value

        if cell_value:  # If there's any value marked (Y, X, date, etc.)
            # Get the date from the header (row 6)
            header_date = ws.cell(row=6, column=col).value

            if header_date and isinstance(header_date, datetime):
                # Keep track of most recent date
                if max_date is None or header_date > max_date:
                    max_date = header_date

    return parse_date(max_date) if max_date else None

def extract_tracking_dates():
    """Extract all proposal tracking dates from Excel"""
    print("=" * 100)
    print("IMPORTING PROPOSAL TRACKING DATES FROM EXCEL")
    print("=" * 100)
    print(f"Excel file: {EXCEL_FILE}")
    print()

    # Load Excel
    print("📄 Loading Excel file...")
    wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
    ws = wb['Weekly proposal']

    print(f"   Dimensions: {ws.dimensions}")
    print(f"   Max row: {ws.max_row}")
    print(f"   Max column: {ws.max_column}")
    print()

    # Extract tracking data
    tracking_data = []
    current_row = 7  # Start at first data row
    projects_found = 0

    print("📊 Extracting tracking dates...")

    while current_row <= ws.max_row:
        # Get project code from column B
        project_code_cell = ws.cell(row=current_row, column=2).value

        if not project_code_cell:
            current_row += 7  # Skip to next project
            continue

        project_code = extract_project_code(project_code_cell)

        if not project_code:
            current_row += 7
            continue

        # Extract dates from calendar structure
        # Row structure: project_row, first_contact, drafting, proposal_sent, contract_signed, on_hold
        first_contact_date = get_most_recent_marked_date(ws, current_row + 1)  # Row 8 for first project
        drafting_date = get_most_recent_marked_date(ws, current_row + 2)
        proposal_sent_date = get_most_recent_marked_date(ws, current_row + 3)
        contract_signed_date = get_most_recent_marked_date(ws, current_row + 4)

        # Calculate most recent activity
        all_dates = [d for d in [first_contact_date, drafting_date, proposal_sent_date, contract_signed_date] if d]
        last_activity_date = max(all_dates) if all_dates else None

        tracking_data.append({
            'project_code': project_code,
            'first_contact_date': first_contact_date,
            'drafting_date': drafting_date,
            'proposal_sent_date': proposal_sent_date,
            'contract_signed_date': contract_signed_date,
            'last_activity_date': last_activity_date
        })

        projects_found += 1

        if projects_found % 10 == 0:
            print(f"   Processed {projects_found} projects...")

        # Move to next project (each project takes 7 rows)
        current_row += 7

    print(f"✅ Extracted tracking data for {projects_found} projects")
    print()

    return tracking_data

def update_database(tracking_data):
    """Update projects table with tracking dates"""
    print("📥 Updating database...")

    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()

    updated_count = 0
    not_found_count = 0

    for data in tracking_data:
        project_code = data['project_code']

        # Check if project exists
        cursor.execute("SELECT project_id FROM projects WHERE project_code = ?", (project_code,))
        result = cursor.fetchone()

        if not result:
            not_found_count += 1
            continue

        project_id = result[0]

        # Update tracking dates with provenance
        cursor.execute("""
            UPDATE projects
            SET
                first_contact_date = ?,
                drafting_date = ?,
                proposal_sent_date = ?,
                contract_signed_date = ?,
                last_proposal_activity_date = ?,
                source_type = ?,
                source_reference = ?,
                updated_by = ?
            WHERE project_id = ?
        """, (
            data['first_contact_date'],
            data['drafting_date'],
            data['proposal_sent_date'],
            data['contract_signed_date'],
            data['last_activity_date'],
            'import',
            'Proposals.xlsx:Weekly_proposal_calendar',
            'import_proposal_tracking_dates',
            project_id
        ))

        updated_count += 1

    conn.commit()

    print(f"   ✅ Updated {updated_count} projects")
    print(f"   ⚠️  {not_found_count} project codes not found in database")
    print()

    # Show some samples
    print("📊 Sample of imported dates:")
    cursor.execute("""
        SELECT
            project_code,
            first_contact_date,
            drafting_date,
            proposal_sent_date,
            contract_signed_date
        FROM projects
        WHERE first_contact_date IS NOT NULL
           OR drafting_date IS NOT NULL
           OR proposal_sent_date IS NOT NULL
        ORDER BY project_code
        LIMIT 10
    """)

    for code, first, draft, sent, signed in cursor.fetchall():
        print(f"   {code}: Contact={first or 'N/A'}, Draft={draft or 'N/A'}, Sent={sent or 'N/A'}, Signed={signed or 'N/A'}")

    conn.close()

def verify_import():
    """Verify import success"""
    print()
    print("=" * 100)
    print("VERIFICATION")
    print("=" * 100)

    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()

    # Count projects with tracking data
    cursor.execute("""
        SELECT
            COUNT(*) as total,
            COUNT(first_contact_date) as has_first_contact,
            COUNT(drafting_date) as has_drafting,
            COUNT(proposal_sent_date) as has_proposal_sent,
            COUNT(contract_signed_date) as has_contract_signed
        FROM projects
    """)

    total, fc, draft, sent, signed = cursor.fetchone()

    print(f"Total projects: {total}")
    print(f"  With first contact date: {fc} ({fc*100/total:.1f}%)")
    print(f"  With drafting date: {draft} ({draft*100/total:.1f}%)")
    print(f"  With proposal sent date: {sent} ({sent*100/total:.1f}%)")
    print(f"  With contract signed date: {signed} ({signed*100/total:.1f}%)")
    print()

    # Show proposals with tracking
    cursor.execute("""
        SELECT
            project_code,
            project_title,
            first_contact_date,
            proposal_sent_date,
            project_stage
        FROM projects
        WHERE project_stage = 'proposal'
        AND (first_contact_date IS NOT NULL OR proposal_sent_date IS NOT NULL)
        ORDER BY proposal_sent_date DESC
        LIMIT 10
    """)

    print("Recent proposals with tracking data:")
    for code, title, contact, sent, stage in cursor.fetchall():
        title_short = title[:50] if title else "N/A"
        print(f"  {code}: {title_short:50} | Contact: {contact or 'N/A'} | Sent: {sent or 'N/A'}")

    conn.close()

    print()
    print("=" * 100)
    print("✅ IMPORT COMPLETE!")
    print("=" * 100)

def main():
    # Extract dates from Excel
    tracking_data = extract_tracking_dates()

    # Update database
    if tracking_data:
        update_database(tracking_data)
        verify_import()
    else:
        print("⚠️  No tracking data extracted")

if __name__ == "__main__":
    main()
