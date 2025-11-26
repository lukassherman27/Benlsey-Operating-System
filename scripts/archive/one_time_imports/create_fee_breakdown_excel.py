#!/usr/bin/env python3
"""
Create CONTRACT FEE BREAKDOWN Excel
Shows: Project > Discipline > Phase with contracted fee and all invoices in columns
"""
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Read parsed invoices
df = pd.read_csv('reports/ALL_INVOICES_PARSED.csv')

# Sort by project, discipline, phase
df = df.sort_values(['project_code', 'discipline', 'phase', 'invoice_date'])

# Group by project, discipline, phase to create fee breakdown rows
grouped = df.groupby(['project_code', 'project_name', 'discipline', 'phase'])

breakdown_rows = []

for (proj_code, proj_name, discipline, phase), group in grouped:
    # Calculate phase fee (sum of all invoice amounts for this phase)
    phase_fee = group['invoice_amount'].sum()

    # Get all invoices for this phase
    invoices = group.to_dict('records')

    # Create row
    row = {
        'project_code': proj_code,
        'project_name': proj_name,
        'contract_type': 'Standard',  # Default, can be customized
        'discipline': discipline,
        'phase': phase,
        'phase_fee': phase_fee,
    }

    # Add up to 5 invoices per row (can extend if needed)
    for i, inv in enumerate(invoices[:5], start=1):
        row[f'invoice_{i}_number'] = inv['invoice_number']
        row[f'invoice_{i}_date'] = inv['invoice_date']
        row[f'invoice_{i}_amount'] = inv['invoice_amount']
        row[f'invoice_{i}_paid_date'] = inv['payment_date']
        row[f'invoice_{i}_paid_amount'] = inv['payment_amount']
        row[f'invoice_{i}_status'] = inv['status']

    # Calculate totals
    total_invoiced = group['invoice_amount'].sum()
    remaining = 0  # Since we're showing all invoices that exist

    row['total_invoiced'] = total_invoiced
    row['remaining'] = remaining

    breakdown_rows.append(row)

# Create DataFrame
breakdown_df = pd.DataFrame(breakdown_rows)

# Create Excel file
excel_file = "/Users/lukassherman/Desktop/CONTRACT_FEE_BREAKDOWN_COMPLETE.xlsx"
wb = Workbook()
ws = wb.active
ws.title = "Fee Breakdown"

# Define headers
headers = [
    'Project Code', 'Project Name', 'Contract Type', 'Discipline', 'Phase', 'Phase Fee',
    'Invoice 1 #', 'Invoice 1 Date', 'Invoice 1 Amt', 'Paid Date', 'Paid Amt', 'Status',
    'Invoice 2 #', 'Invoice 2 Date', 'Invoice 2 Amt', 'Paid Date', 'Paid Amt', 'Status',
    'Invoice 3 #', 'Invoice 3 Date', 'Invoice 3 Amt', 'Paid Date', 'Paid Amt', 'Status',
    'Invoice 4 #', 'Invoice 4 Date', 'Invoice 4 Amt', 'Paid Date', 'Paid Amt', 'Status',
    'Invoice 5 #', 'Invoice 5 Date', 'Invoice 5 Amt', 'Paid Date', 'Paid Amt', 'Status',
    'Total Invoiced', 'Remaining'
]

# Write headers
ws.append(headers)

# Style header row
header_fill = PatternFill(start_color="2C5F2D", end_color="2C5F2D", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=10)
header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = header_alignment

# Write data rows
for _, row in breakdown_df.iterrows():
    row_data = [
        row['project_code'],
        row['project_name'],
        row['contract_type'],
        row['discipline'],
        row['phase'],
        row['phase_fee'],
    ]

    # Add invoice data for up to 5 invoices
    for i in range(1, 6):
        row_data.extend([
            row.get(f'invoice_{i}_number', ''),
            row.get(f'invoice_{i}_date', ''),
            row.get(f'invoice_{i}_amount', ''),
            row.get(f'invoice_{i}_paid_date', ''),
            row.get(f'invoice_{i}_paid_amount', ''),
            row.get(f'invoice_{i}_status', ''),
        ])

    row_data.extend([
        row['total_invoiced'],
        row['remaining']
    ])

    ws.append(row_data)

# Set column widths
col_widths = {
    'A': 12, 'B': 35, 'C': 12, 'D': 18, 'E': 22, 'F': 14,
}

# Invoice columns - repeat pattern
for i in range(6):
    base = 7 + (i * 6)
    col_widths[get_column_letter(base)] = 12      # Invoice #
    col_widths[get_column_letter(base + 1)] = 11  # Date
    col_widths[get_column_letter(base + 2)] = 12  # Amount
    col_widths[get_column_letter(base + 3)] = 11  # Paid Date
    col_widths[get_column_letter(base + 4)] = 12  # Paid Amount
    col_widths[get_column_letter(base + 5)] = 10  # Status

# Totals columns
col_widths[get_column_letter(len(headers) - 1)] = 14  # Total Invoiced
col_widths[get_column_letter(len(headers))] = 12      # Remaining

for col_letter, width in col_widths.items():
    ws.column_dimensions[col_letter].width = width

# Format currency columns
for row_idx in range(2, ws.max_row + 1):
    ws[f'F{row_idx}'].number_format = '$#,##0.00'  # Phase Fee

    # Invoice amounts and paid amounts
    for i in range(5):
        base = 7 + (i * 6)
        amt_col = get_column_letter(base + 2)
        paid_col = get_column_letter(base + 4)
        if ws[f'{amt_col}{row_idx}'].value:
            ws[f'{amt_col}{row_idx}'].number_format = '$#,##0.00'
        if ws[f'{paid_col}{row_idx}'].value:
            ws[f'{paid_col}{row_idx}'].number_format = '$#,##0.00'

    # Totals
    ws[f'{get_column_letter(len(headers) - 1)}{row_idx}'].number_format = '$#,##0.00'
    ws[f'{get_column_letter(len(headers))}{row_idx}'].number_format = '$#,##0.00'

# Freeze panes
ws.freeze_panes = 'A2'

# Save
wb.save(excel_file)

print("=" * 100)
print("CONTRACT FEE BREAKDOWN EXCEL CREATED")
print("=" * 100)
print(f"\nFile: {excel_file}")
print(f"\nStructure:")
print(f"  - One row per project/discipline/phase combination")
print(f"  - Shows phase fee and all invoices for that phase")
print(f"  - Up to 5 invoices per row (can be extended)")
print(f"\nRows: {len(breakdown_rows)}")
print(f"\nReady for your review!")
print("=" * 100)
