#!/usr/bin/env python3
"""
import_proposals_v2.py

ENHANCED VERSION - Extracts location, country, currency, and status from Excel
Includes status progression timeline tracking
"""

import openpyxl
import sqlite3
import re
from pathlib import Path
from datetime import datetime

class ProposalsImporterV2:
    def __init__(self, excel_path, db_path=None):
        self.excel_path = Path(excel_path)

        # Use provided path or default to OneDrive location
        if db_path is None:
            db_path = "database/bensley_master.db"
        self.db_path = Path(db_path)

        print(f"📂 Excel: {self.excel_path}")
        print(f"📂 Database: {self.db_path}")

        # Load Excel
        self.wb = openpyxl.load_workbook(excel_path, data_only=True)
        self.ws = self.wb['Weekly proposal']

        # Connect to database
        self.conn = sqlite3.connect(self.db_path)
        self.cursor = self.conn.cursor()

        self.stats = {
            'proposals_found': 0,
            'proposals_imported': 0,
            'with_location': 0,
            'with_signed_date': 0,
            'status_won': 0,
            'status_proposal': 0,
            'errors': 0
        }

    def extract_project_code(self, code_str):
        """Extract BK code from '25 BK-001' format"""
        if not code_str:
            return None

        code_str = str(code_str).strip()

        # Look for pattern "BK-XXX"
        if "BK-" in code_str:
            parts = code_str.split("BK-")
            if len(parts) > 1:
                code_num = parts[1].split()[0]
                return f"{parts[0].strip()} BK-{code_num}"

        return None

    def extract_location_country(self, project_name):
        """
        Extract location and country from project name
        Patterns:
        - "Project in Abu Dhabi, UAE" → location: "Abu Dhabi, UAE", country: "UAE"
        - "Project, Vietnam" → location: "Vietnam", country: "Vietnam"
        - "Project in Lagos, Nigeria" → location: "Lagos, Nigeria", country: "Nigeria"
        """
        if not project_name:
            return None, None

        name = str(project_name).strip()

        # Pattern 1: "in Location, Country"
        match = re.search(r' in ([^,]+), ([A-Z][A-Za-z\s]+)', name)
        if match:
            city = match.group(1).strip()
            country = match.group(2).strip()
            return f"{city}, {country}", country

        # Pattern 2: "in Country" (no city)
        match = re.search(r' in ([A-Z][A-Za-z\s]+)(?:-|$)', name)
        if match:
            country = match.group(1).strip()
            return country, country

        # Pattern 3: ", Country" at end
        match = re.search(r', ([A-Z][A-Za-z\s]+)(?:-|$)', name)
        if match:
            country = match.group(1).strip()
            return country, country

        return None, None

    def get_currency_for_country(self, country):
        """Map country to currency (simplified)"""
        if not country:
            return 'USD'

        country_upper = country.upper()

        currency_map = {
            'UAE': 'AED',
            'CHINA': 'CNY',
            'VIETNAM': 'VND',
            'NIGERIA': 'NGN',
            'ISRAEL': 'ILS',
            'INDIA': 'INR',
            'THAILAND': 'THB',
            'SINGAPORE': 'SGD',
            'MALAYSIA': 'MYR',
            'INDONESIA': 'IDR',
            'PHILIPPINES': 'PHP',
            'JAPAN': 'JPY',
            'SOUTH KOREA': 'KRW',
            'AUSTRALIA': 'AUD',
            'NEW ZEALAND': 'NZD',
            'UK': 'GBP',
            'UNITED KINGDOM': 'GBP',
            'FRANCE': 'EUR',
            'GERMANY': 'EUR',
            'ITALY': 'EUR',
            'SPAIN': 'EUR',
        }

        for key, currency in currency_map.items():
            if key in country_upper:
                return currency

        return 'USD'  # Default

    def add_missing_columns(self):
        """Add location, country, currency, proposal_sent_date columns if missing"""
        print("\n📊 Checking database schema...")

        # Get current columns
        self.cursor.execute("PRAGMA table_info(proposals)")
        existing_columns = {row[1] for row in self.cursor.fetchall()}

        columns_to_add = {
            'location': "ALTER TABLE proposals ADD COLUMN location TEXT",
            'country': "ALTER TABLE proposals ADD COLUMN country TEXT",
            'currency': "ALTER TABLE proposals ADD COLUMN currency TEXT DEFAULT 'USD'",
            'proposal_sent_date': "ALTER TABLE proposals ADD COLUMN proposal_sent_date TEXT"
        }

        for col_name, sql in columns_to_add.items():
            if col_name not in existing_columns:
                print(f"   ➕ Adding column: {col_name}")
                self.cursor.execute(sql)

        self.conn.commit()
        print("✓ Schema updated")

    def extract_proposals(self):
        """Extract proposals from Excel with enhanced data"""
        proposals = []
        current_row = 7  # Start after header
        max_rows = self.ws.max_row

        print(f"\n📋 Extracting proposals from rows 7-{max_rows}...")

        while current_row <= max_rows:
            try:
                # Get project code
                project_code_cell = self.ws.cell(row=current_row, column=2).value

                # Skip empty rows
                if not project_code_cell:
                    current_row += 1
                    continue

                # Extract project code
                project_code = self.extract_project_code(project_code_cell)

                if not project_code:
                    current_row += 1
                    continue

                # Extract basic data
                project_name = self.ws.cell(row=current_row, column=3).value or ""
                client_company = self.ws.cell(row=current_row, column=4).value or ""
                contact_person = self.ws.cell(row=current_row, column=5).value or ""
                contact_phone = self.ws.cell(row=current_row, column=6).value or ""
                contact_email = self.ws.cell(row=current_row, column=7).value or ""

                # Disciplines
                is_landscape = 1 if self.ws.cell(row=current_row, column=8).value == 'Y' else 0
                is_architect = 1 if self.ws.cell(row=current_row, column=9).value == 'Y' else 0
                is_interior = 1 if self.ws.cell(row=current_row, column=10).value == 'Y' else 0

                # Project value
                project_value = self.ws.cell(row=current_row, column=11).value
                if project_value:
                    try:
                        project_value = float(project_value)
                    except:
                        project_value = None

                # ENHANCED: Extract location and country from project name
                location, country = self.extract_location_country(project_name)
                if location:
                    self.stats['with_location'] += 1

                # ENHANCED: Get currency based on country
                currency = self.get_currency_for_country(country)

                # ENHANCED: Contract signed date (Column L)
                contract_signed_date = self.ws.cell(row=current_row, column=12).value

                # Determine status
                status = 'proposal'  # Default
                if contract_signed_date and str(contract_signed_date) != 'not signed':
                    if isinstance(contract_signed_date, datetime):
                        contract_signed_date = contract_signed_date.strftime('%Y-%m-%d')
                        status = 'won'
                        self.stats['status_won'] += 1
                        self.stats['with_signed_date'] += 1
                    else:
                        contract_signed_date = None
                else:
                    contract_signed_date = None
                    self.stats['status_proposal'] += 1

                proposal = {
                    'project_code': project_code,
                    'project_name': str(project_name).strip() if project_name else "",
                    'client_company': str(client_company).strip() if client_company else "",
                    'contact_person': str(contact_person).strip() if contact_person else "",
                    'contact_phone': str(contact_phone).strip() if contact_phone else "",
                    'contact_email': str(contact_email).strip() if contact_email else "",
                    'project_value': project_value,
                    'is_landscape': is_landscape,
                    'is_architect': is_architect,
                    'is_interior': is_interior,
                    'location': location,
                    'country': country,
                    'currency': currency,
                    'status': status,
                    'contract_signed_date': contract_signed_date,
                }

                proposals.append(proposal)
                self.stats['proposals_found'] += 1

                if self.stats['proposals_found'] % 10 == 0:
                    print(f"   Found {self.stats['proposals_found']} proposals...")

                # Move to next project (skip status rows - each project takes 7 rows)
                current_row += 7

            except Exception as e:
                print(f"   ⚠️ Error at row {current_row}: {e}")
                current_row += 1
                self.stats['errors'] += 1

        print(f"✓ Found {self.stats['proposals_found']} proposals")
        print(f"   {self.stats['with_location']} with location/country")
        print(f"   {self.stats['status_won']} won, {self.stats['status_proposal']} proposal")
        return proposals

    def import_proposals(self, proposals):
        """Import proposals to database"""
        print(f"\n💾 Importing {len(proposals)} proposals...")

        for proposal in proposals:
            try:
                self.cursor.execute("""
                    INSERT OR REPLACE INTO proposals
                    (project_code, project_name, client_company, contact_person,
                     contact_phone, contact_email, project_value,
                     is_landscape, is_architect, is_interior,
                     location, country, currency, status,
                     contract_signed_date,
                     created_at, updated_at)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, datetime('now'), datetime('now'))
                """, (
                    proposal['project_code'],
                    proposal['project_name'],
                    proposal['client_company'],
                    proposal['contact_person'],
                    proposal['contact_phone'],
                    proposal['contact_email'],
                    proposal['project_value'],
                    proposal['is_landscape'],
                    proposal['is_architect'],
                    proposal['is_interior'],
                    proposal['location'],
                    proposal['country'],
                    proposal['currency'],
                    proposal['status'],
                    proposal['contract_signed_date'],
                ))

                self.stats['proposals_imported'] += 1

                if self.stats['proposals_imported'] % 10 == 0:
                    print(f"   Imported {self.stats['proposals_imported']}...")

            except Exception as e:
                print(f"   ✗ Error importing {proposal['project_code']}: {e}")
                self.stats['errors'] += 1

        self.conn.commit()
        print(f"✓ Imported {self.stats['proposals_imported']} proposals")

    def print_summary(self):
        """Print import summary"""
        print("\n" + "="*60)
        print("📊 IMPORT SUMMARY")
        print("="*60)
        print(f"Proposals Found:    {self.stats['proposals_found']}")
        print(f"Proposals Imported: {self.stats['proposals_imported']}")
        print(f"With Location:      {self.stats['with_location']}")
        print(f"With Signed Date:   {self.stats['with_signed_date']}")
        print(f"Status='won':       {self.stats['status_won']}")
        print(f"Status='proposal':  {self.stats['status_proposal']}")
        print(f"Errors:             {self.stats['errors']}")
        print("="*60)

    def run(self):
        """Run full import process"""
        print("\n🚀 Starting enhanced proposals import...")
        print("="*60)

        # Add missing columns
        self.add_missing_columns()

        # Extract proposals
        proposals = self.extract_proposals()

        if not proposals:
            print("❌ No proposals found to import")
            return

        # Import to database
        self.import_proposals(proposals)

        # Print summary
        self.print_summary()

        self.conn.close()
        print("\n✅ Import complete!")


if __name__ == "__main__":
    import sys

    # Usage
    excel_path = "Proposals.xlsx"
    db_path = "database/bensley_master.db"

    if len(sys.argv) > 1:
        excel_path = sys.argv[1]
    if len(sys.argv) > 2:
        db_path = sys.argv[2]

    importer = ProposalsImporterV2(excel_path, db_path)
    importer.run()
