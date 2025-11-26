#!/usr/bin/env python3
"""
import_proposal_dashboard.py

Imports proposal tracking data from "Proposal dashboard " sheet
Includes: status progression, timeline data, days in each phase
"""

import openpyxl
import sqlite3
from pathlib import Path
from datetime import datetime

class ProposalDashboardImporter:
    def __init__(self, excel_path, db_path=None):
        self.excel_path = Path(excel_path)

        if db_path is None:
            db_path = "database/bensley_master.db"
        self.db_path = Path(db_path)

        print(f"📂 Excel: {self.excel_path}")
        print(f"📂 Database: {self.db_path}")

        # Load Excel - note the trailing space in sheet name!
        self.wb = openpyxl.load_workbook(excel_path, data_only=True)
        self.ws = self.wb['Proposal dashboard ']  # Trailing space is important!

        # Connect to database
        self.conn = sqlite3.connect(self.db_path)
        self.cursor = self.conn.cursor()

        self.stats = {
            'proposals_found': 0,
            'proposals_updated': 0,
            'with_status': 0,
            'with_dates': 0,
            'errors': 0
        }

    def parse_date(self, date_val):
        """Parse date from Excel"""
        if not date_val:
            return None
        if isinstance(date_val, datetime):
            return date_val.strftime('%Y-%m-%d')
        if isinstance(date_val, str):
            # Try parsing string dates
            for fmt in ['%Y-%m-%d', '%m/%d/%Y', '%d/%m/%Y']:
                try:
                    return datetime.strptime(date_val.strip(), fmt).strftime('%Y-%m-%d')
                except:
                    continue
        return None

    def extract_project_code(self, code_str):
        """Extract project code"""
        if not code_str:
            return None
        return str(code_str).strip()

    def extract_proposals(self):
        """Extract proposals from dashboard sheet"""
        proposals = []
        current_row = 7  # Start after header (row 6)
        max_rows = self.ws.max_row

        print(f"\n📋 Extracting proposals from rows 7-{max_rows}...")

        while current_row <= max_rows:
            try:
                # Get project code (column 2)
                project_code_cell = self.ws.cell(current_row, 2).value

                # Skip empty rows
                if not project_code_cell:
                    current_row += 1
                    continue

                project_code = self.extract_project_code(project_code_cell)
                if not project_code or not project_code.startswith('25 BK-'):
                    current_row += 1
                    continue

                # Extract data from dashboard
                project_name = self.ws.cell(current_row, 3).value
                country = self.ws.cell(current_row, 4).value

                # Status tracking
                last_week_status = self.ws.cell(current_row, 13).value
                current_status = self.ws.cell(current_row, 15).value
                days_in_status = self.ws.cell(current_row, 17).value

                # Dates
                proposal_sent = self.ws.cell(current_row, 19).value  # Y/N
                contract_signed_date = self.parse_date(self.ws.cell(current_row, 20).value)
                first_contact_date = self.parse_date(self.ws.cell(current_row, 21).value)
                days_to_sign = self.ws.cell(current_row, 22).value
                proposal_sent_date = self.parse_date(self.ws.cell(current_row, 23).value)

                # Workflow metrics
                num_proposals = self.ws.cell(current_row, 24).value
                days_in_drafting = self.ws.cell(current_row, 25).value
                days_in_review = self.ws.cell(current_row, 26).value

                # Other
                remarks = self.ws.cell(current_row, 28).value
                phase = self.ws.cell(current_row, 30).value

                proposal = {
                    'project_code': project_code,
                    'project_name': str(project_name).strip() if project_name else None,
                    'country': str(country).strip() if country else None,
                    'last_week_status': str(last_week_status).strip() if last_week_status else None,
                    'current_status': str(current_status).strip() if current_status else 'first_contact',
                    'days_in_status': int(days_in_status) if days_in_status and str(days_in_status).isdigit() else None,
                    'proposal_sent': str(proposal_sent).upper() == 'Y' if proposal_sent else False,
                    'contract_signed_date': contract_signed_date,
                    'first_contact_date': first_contact_date,
                    'days_to_sign': int(days_to_sign) if days_to_sign and str(days_to_sign).replace('.', '').isdigit() else None,
                    'proposal_sent_date': proposal_sent_date,
                    'num_proposals': int(num_proposals) if num_proposals and str(num_proposals).isdigit() else 1,
                    'days_in_drafting': int(days_in_drafting) if days_in_drafting and str(days_in_drafting).replace('.', '').isdigit() else None,
                    'days_in_review': int(days_in_review) if days_in_review and str(days_in_review).replace('.', '').isdigit() else None,
                    'remarks': str(remarks).strip() if remarks else None,
                    'phase': str(phase).strip() if phase else None,
                }

                proposals.append(proposal)
                self.stats['proposals_found'] += 1

                if current_status:
                    self.stats['with_status'] += 1
                if first_contact_date or proposal_sent_date:
                    self.stats['with_dates'] += 1

                if self.stats['proposals_found'] % 10 == 0:
                    print(f"   Found {self.stats['proposals_found']} proposals...")

                # Each proposal takes 7 rows in the dashboard
                current_row += 7

            except Exception as e:
                print(f"   ⚠️ Error at row {current_row}: {e}")
                current_row += 1
                self.stats['errors'] += 1

        print(f"✓ Found {self.stats['proposals_found']} proposals")
        print(f"   {self.stats['with_status']} with current status")
        print(f"   {self.stats['with_dates']} with timeline dates")
        return proposals

    def add_missing_columns(self):
        """Add timeline tracking columns if missing"""
        print("\n📊 Checking database schema...")

        # Get current columns
        self.cursor.execute("PRAGMA table_info(proposals)")
        existing_columns = {row[1] for row in self.cursor.fetchall()}

        columns_to_add = {
            'current_status': "ALTER TABLE proposals ADD COLUMN current_status TEXT",
            'last_week_status': "ALTER TABLE proposals ADD COLUMN last_week_status TEXT",
            'days_in_current_status': "ALTER TABLE proposals ADD COLUMN days_in_current_status INTEGER",
            'first_contact_date': "ALTER TABLE proposals ADD COLUMN first_contact_date DATE",
            'proposal_sent_date': "ALTER TABLE proposals ADD COLUMN proposal_sent_date DATE",
            'days_to_sign': "ALTER TABLE proposals ADD COLUMN days_to_sign INTEGER",
            'days_in_drafting': "ALTER TABLE proposals ADD COLUMN days_in_drafting INTEGER",
            'days_in_review': "ALTER TABLE proposals ADD COLUMN days_in_review INTEGER",
            'num_proposals_sent': "ALTER TABLE proposals ADD COLUMN num_proposals_sent INTEGER",
            'phase': "ALTER TABLE proposals ADD COLUMN phase TEXT",
            'remarks': "ALTER TABLE proposals ADD COLUMN remarks TEXT",
        }

        for col_name, sql in columns_to_add.items():
            if col_name not in existing_columns:
                print(f"   ➕ Adding column: {col_name}")
                self.cursor.execute(sql)

        self.conn.commit()
        print("✓ Schema updated")

    def update_proposals(self, proposals):
        """Update proposals in database"""
        print(f"\n💾 Updating {len(proposals)} proposals...")

        for proposal in proposals:
            try:
                self.cursor.execute("""
                    UPDATE proposals
                    SET
                        current_status = ?,
                        last_week_status = ?,
                        days_in_current_status = ?,
                        first_contact_date = ?,
                        proposal_sent_date = ?,
                        days_to_sign = ?,
                        days_in_drafting = ?,
                        days_in_review = ?,
                        num_proposals_sent = ?,
                        phase = ?,
                        remarks = ?,
                        updated_at = datetime('now')
                    WHERE project_code = ?
                """, (
                    proposal['current_status'],
                    proposal['last_week_status'],
                    proposal['days_in_status'],
                    proposal['first_contact_date'],
                    proposal['proposal_sent_date'],
                    proposal['days_to_sign'],
                    proposal['days_in_drafting'],
                    proposal['days_in_review'],
                    proposal['num_proposals'],
                    proposal['phase'],
                    proposal['remarks'],
                    proposal['project_code'],
                ))

                if self.cursor.rowcount > 0:
                    self.stats['proposals_updated'] += 1

                if self.stats['proposals_updated'] % 10 == 0:
                    print(f"   Updated {self.stats['proposals_updated']}...")

            except Exception as e:
                print(f"   ✗ Error updating {proposal['project_code']}: {e}")
                self.stats['errors'] += 1

        self.conn.commit()
        print(f"✓ Updated {self.stats['proposals_updated']} proposals")

    def print_summary(self):
        """Print import summary"""
        print("\n" + "="*60)
        print("📊 IMPORT SUMMARY")
        print("="*60)
        print(f"Proposals Found:    {self.stats['proposals_found']}")
        print(f"Proposals Updated:  {self.stats['proposals_updated']}")
        print(f"With Status:        {self.stats['with_status']}")
        print(f"With Timeline Data: {self.stats['with_dates']}")
        print(f"Errors:             {self.stats['errors']}")
        print("="*60)

    def run(self):
        """Run full import process"""
        print("\n🚀 Starting proposal dashboard import...")
        print("="*60)

        # Add missing columns
        self.add_missing_columns()

        # Extract proposals
        proposals = self.extract_proposals()

        if not proposals:
            print("❌ No proposals found to import")
            return

        # Update database
        self.update_proposals(proposals)

        # Print summary
        self.print_summary()

        self.conn.close()
        print("\n✅ Import complete!")


if __name__ == "__main__":
    import sys

    excel_path = "Proposals.xlsx"
    db_path = "database/bensley_master.db"

    if len(sys.argv) > 1:
        excel_path = sys.argv[1]
    if len(sys.argv) > 2:
        db_path = sys.argv[2]

    importer = ProposalDashboardImporter(excel_path, db_path)
    importer.run()
